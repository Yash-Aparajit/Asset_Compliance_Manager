import pandas as pd
import os
import zipfile

from flask import (
    Flask, render_template, request,
    redirect, url_for, session, flash, send_file, jsonify
)

from models import (
    db, User, Asset, AMC, AMCEvent, AMCDocument,
    Calibration, CalibrationEvent, CalibrationDocument,
    AssetScrap, ReminderAck
)

from datetime import datetime, date
from werkzeug.security import generate_password_hash, check_password_hash
from functools import wraps

from openpyxl import Workbook
from io import BytesIO


# -------------------------------------------------
# APP CONFIG
# -------------------------------------------------
app = Flask(__name__)
app.config["SECRET_KEY"] = "acm-secret-key-change-later"

BASE_DIR = os.path.abspath(os.path.dirname(__file__))

DATA_DIR = os.path.join(BASE_DIR, "data")
os.makedirs(DATA_DIR, exist_ok=True)

DB_PATH = os.path.join(DATA_DIR, "app.db")

app.config["SQLALCHEMY_DATABASE_URI"] = f"sqlite:///{DB_PATH}"
app.config["SQLALCHEMY_TRACK_MODIFICATIONS"] = False

app.config["SQLALCHEMY_ENGINE_OPTIONS"] = {
    "connect_args": {"timeout": 15}
}

db.init_app(app)

# -------------------------------------------------
# FILE STORAGE PATHS
# -------------------------------------------------
AMC_DOC_DIR = os.path.join(DATA_DIR, "AMC")
os.makedirs(AMC_DOC_DIR, exist_ok=True)

CALIBRATION_DOC_DIR = os.path.join(DATA_DIR, "Calibration")
os.makedirs(CALIBRATION_DOC_DIR, exist_ok=True)

SCRAP_DOC_DIR = os.path.join(DATA_DIR, "Scrap")
os.makedirs(SCRAP_DOC_DIR, exist_ok=True)

# -------------------------------------------------
# Utility Helpers
# -------------------------------------------------
def format_date_indian(date_obj):
    if not date_obj:
        return "-"
    return date_obj.strftime("%d/%m/%Y")


app.jinja_env.globals.update(
    format_date_indian=format_date_indian
)


def parse_indian_date(val):
    if val in (None, "", pd.NaT):
        return None

    if isinstance(val, (datetime, date)):
        return val.date() if isinstance(val, datetime) else val

    val = str(val).strip()

    for fmt in ("%d/%m/%Y", "%d-%m-%Y"):
        try:
            return datetime.strptime(val, fmt).date()
        except Exception:
            continue

    return None


def get_amc_status(amc):
    today = date.today()
    days_left = (amc.end_date - today).days

    if amc.is_completed:
        return ("completed", "Completed")

    if amc.is_cancelled:
        return ("cancelled", "Cancelled")

    if days_left < 0:
        return ("overdue", "Overdue")

    if days_left <= 30:
        return ("warning", "Expiring Soon")

    return ("active", "Active")


def is_acknowledged(source_type, source_id, rule):
    return (
        ReminderAck.query
        .filter_by(
            source_type=source_type,
            source_id=source_id,
            rule=rule
        )
        .first()
        is not None
    )


# -------------------------------------------------
# ASSET IMPORT SCHEMA
# -------------------------------------------------
ASSET_IMPORT_COLUMNS = {
    "Asset Code": "asset_code",
    "Asset Name": "asset_name",
    "Serial No": "serial_no",
    "Plant": "plant",
    "Department": "department",
    "Location": "location",
    "Purchase Date (DD/MM/YYYY)": "purchase_date",
}


# -------------------------------------------------
# AUTH HELPERS
# -------------------------------------------------
def login_required(f):
    @wraps(f)
    def wrapper(*args, **kwargs):
        if "user_id" not in session:
            return redirect(url_for("login"))
        return f(*args, **kwargs)
    return wrapper


def role_required(roles):
    def decorator(f):
        @wraps(f)
        def wrapper(*args, **kwargs):
            if session.get("role") not in roles:
                flash("Access denied", "danger")
                return redirect(url_for("asset_master"))
            return f(*args, **kwargs)
        return wrapper
    return decorator


# -------------------------------------------------
# INITIAL DB + USER SEED
# -------------------------------------------------
def seed_users():
    users = [
        {
            "username": "developer",
            "password": "dev@123@123",
            "role": "developer"
        },
        {
            "username": "purchase",
            "password": "purchase@jeena",
            "role": "purchase"
        }
    ]

    for u in users:
        existing = User.query.filter_by(username=u["username"]).first()
        if not existing:
            user = User(
                username=u["username"],
                password_hash=generate_password_hash(u["password"]),
                role=u["role"]
            )
            db.session.add(user)

    try:
        db.session.commit()
    except Exception:
        db.session.rollback()


with app.app_context():
    db.create_all()
    seed_users()

    # WAL mode reduces sqlite write-lock pain (safe)
    try:
        db.session.execute(db.text("PRAGMA journal_mode=WAL;"))
        db.session.commit()
    except Exception:
        db.session.rollback()


# -------------------------------------------------
# ROOT ROUTE
# -------------------------------------------------
@app.route("/")
def root():
    if "user_id" not in session:
        return redirect(url_for("login"))
    return redirect(url_for("reminders"))


# -------------------------------------------------
# LOGIN & LOGOUT
# -------------------------------------------------
@app.route("/login", methods=["GET", "POST"])
def login():
    if request.method == "POST":
        username = request.form.get("username")
        password = request.form.get("password")

        user = User.query.filter_by(username=username).first()

        if not user or not check_password_hash(user.password_hash, password):
            flash("Invalid credentials", "danger")
            return render_template("login.html")

        session["user_id"] = user.id
        session["username"] = user.username
        session["role"] = user.role

        return redirect(url_for("reminders"))

    return render_template("login.html")


@app.route("/logout")
@login_required
def logout():
    session.clear()
    return redirect(url_for("login"))


# -------------------------------------------------
# ASSET MASTER 
# -------------------------------------------------
@app.route("/assets")
@login_required
def asset_master():
    query = Asset.query

    search = request.args.get("search")
    plant = request.args.get("plant")
    department = request.args.get("department")
    status = request.args.get("status")

    if search:
        query = query.filter(
            (Asset.asset_code.ilike(f"%{search}%")) |
            (Asset.asset_name.ilike(f"%{search}%")) |
            (Asset.serial_no.ilike(f"%{search}%"))
        )

    if plant:
        query = query.filter(Asset.plant == plant)

    if department:
        query = query.filter(Asset.department == department)

    if status:
        query = query.filter(Asset.status == status)

    page = request.args.get("page", 1, type=int)
    per_page = 20

    pagination = query.order_by(Asset.id.asc()).paginate(
        page=page,
        per_page=per_page,
        error_out=False
    )

    return render_template(
        "asset_master.html",
        assets=pagination.items,
        pagination=pagination
    )


# -------------------------------------------------
# ADD ASSET
# -------------------------------------------------
@app.route("/assets/add", methods=["GET", "POST"])
@login_required
def add_asset():
    if request.method == "POST":
        asset_code = request.form.get("asset_code", "").strip()
        asset_name = request.form.get("asset_name", "").strip()

        serial_no = request.form.get("serial_no")
        serial_no = serial_no.strip() if serial_no else None

        if not asset_code or not asset_name:
            flash("Asset Code and Asset Name are required", "danger")
            return render_template("asset_add.html")

        if Asset.query.filter_by(asset_code=asset_code).first():
            flash("Asset Code already exists", "danger")
            return render_template("asset_add.html")

        if serial_no:
            if Asset.query.filter_by(serial_no=serial_no).first():
                flash("Serial No already exists for another asset", "danger")
                return render_template("asset_add.html")

        purchase_date_str = request.form.get("purchase_date")
        if purchase_date_str:
            purchase_date = datetime.strptime(purchase_date_str, "%Y-%m-%d").date()
        else:
            purchase_date = None

        asset = Asset(
            asset_code=asset_code,
            asset_name=asset_name,
            serial_no=serial_no,
            plant=request.form.get("plant"),
            department=request.form.get("department"),
            location=request.form.get("location"),
            purchase_date=purchase_date
        )

        try:
            db.session.add(asset)
            db.session.commit()
            flash("Asset added successfully", "success")
            return redirect(url_for("asset_master"))
        except Exception:
            db.session.rollback()
            flash("Asset add failed. No data was saved.", "danger")
            return render_template("asset_add.html")

    return render_template("asset_add.html")


# -------------------------------------------------
# EDIT ASSET
# -------------------------------------------------
@app.route("/assets/edit/<int:asset_id>", methods=["GET", "POST"])
@login_required
def edit_asset(asset_id):
    asset = Asset.query.get_or_404(asset_id)

    if asset.status == "Scrapped":
        flash("Scrapped assets cannot be edited", "danger")
        return redirect(url_for("asset_master"))

    if request.method == "POST":
        new_code = request.form.get("asset_code", "").strip()
        new_name = request.form.get("asset_name", "").strip()

        serial_no = request.form.get("serial_no")
        serial_no = serial_no.strip() if serial_no else None

        if not new_code or not new_name:
            flash("Asset Code and Asset Name are required", "danger")
            return render_template("asset_edit.html", asset=asset)

        if new_code != asset.asset_code:
            if Asset.query.filter_by(asset_code=new_code).first():
                flash("Asset Code already exists", "danger")
                return render_template("asset_edit.html", asset=asset)

        if serial_no:
            existing_serial = (
                Asset.query
                .filter(Asset.serial_no == serial_no, Asset.id != asset.id)
                .first()
            )
            if existing_serial:
                flash("Serial No already exists for another asset", "danger")
                return render_template("asset_edit.html", asset=asset)

        asset.asset_code = new_code
        asset.asset_name = new_name
        asset.serial_no = serial_no
        asset.plant = request.form.get("plant")
        asset.department = request.form.get("department")
        asset.location = request.form.get("location")

        purchase_date_str = request.form.get("purchase_date")
        if purchase_date_str:
            asset.purchase_date = datetime.strptime(purchase_date_str, "%Y-%m-%d").date()
        else:
            asset.purchase_date = None

        try:
            db.session.commit()
            flash("Asset updated successfully", "success")
            return redirect(url_for("asset_master"))

        except Exception:
            db.session.rollback()
            flash("Asset update failed. No changes were saved.", "danger")
            return render_template("asset_edit.html", asset=asset)

    return render_template("asset_edit.html", asset=asset)


# -------------------------------------------------
# Asset Import
# -------------------------------------------------
@app.route("/assets/import", methods=["GET", "POST"])
@login_required
def import_assets():
    if request.method == "POST":
        file = request.files.get("file")

        if not file or not file.filename.lower().endswith(".xlsx"):
            flash("Please upload a valid .xlsx Excel file", "danger")
            return redirect(url_for("import_assets"))

        try:
            df = pd.read_excel(file)
        except Exception:
            flash("Unable to read Excel file. File may be corrupted.", "danger")
            return redirect(url_for("import_assets"))

        if df.empty:
            flash(
                "Excel file contains no data rows. Please add asset data below the header.",
                "warning"
            )
            return redirect(url_for("import_assets"))

        uploaded_columns = [str(c).strip() for c in df.columns]
        expected_columns = list(ASSET_IMPORT_COLUMNS.keys())

        missing = set(expected_columns) - set(uploaded_columns)
        if missing:
            flash(f"Missing columns: {', '.join(missing)}", "danger")
            return redirect(url_for("import_assets"))

        df = df.rename(columns=ASSET_IMPORT_COLUMNS)

        valid_rows = []
        invalid_rows = []

        seen_codes = set()
        existing_codes = {
            a.asset_code
            for a in Asset.query.with_entities(Asset.asset_code).all()
        }

        existing_serials = {
            a.serial_no
            for a in Asset.query.with_entities(Asset.serial_no).all()
            if a.serial_no
        }

        for idx, row in df.iterrows():
            if row.isna().all():
                continue

            errors = []

            asset_code = row.get("asset_code")
            asset_name = row.get("asset_name")

            asset_code = str(asset_code).strip() if pd.notna(asset_code) else ""
            asset_name = str(asset_name).strip() if pd.notna(asset_name) else ""

            serial_no = row.get("serial_no")
            serial_no = str(serial_no).strip() if pd.notna(serial_no) else ""

            if not asset_code:
                errors.append("Missing Asset Code")
            if not asset_name:
                errors.append("Missing Asset Name")

            if asset_code:
                if asset_code in seen_codes:
                    errors.append("Duplicate Asset Code in file")
                if asset_code in existing_codes:
                    errors.append("Asset Code already exists")

            if asset_code:
                seen_codes.add(asset_code)

            if serial_no:
                if serial_no in existing_serials:
                    errors.append("Serial No already exists")

            purchase_date = None
            raw_date = row.get("purchase_date")

            if raw_date not in (None, "", pd.NaT):
                parsed_date = parse_indian_date(raw_date)
                purchase_date = parsed_date.isoformat() if parsed_date else None

                if not purchase_date:
                    errors.append("Invalid Purchase Date (DD/MM/YYYY or DD-MM-YYYY)")

                if parsed_date and parsed_date > date.today():
                    errors.append("Purchase Date cannot be in the future")

            record = {
                "row_no": idx + 2,
                "asset_code": asset_code,
                "asset_name": asset_name,
                "serial_no": serial_no if serial_no else None,
                "plant": row.get("plant"),
                "department": row.get("department"),
                "location": row.get("location"),
                "purchase_date": purchase_date,
            }

            if errors:
                record["errors"] = errors
                invalid_rows.append(record)
            else:
                valid_rows.append(record)

        if not valid_rows and not invalid_rows:
            flash(
                "All rows in the file are empty. Please enter asset details and try again.",
                "warning"
            )
            return redirect(url_for("import_assets"))

        if not valid_rows and invalid_rows:
            flash(
                "All rows are invalid. Please fix errors and re-upload the file.",
                "danger"
            )
            return redirect(url_for("import_assets"))

        session["import_valid"] = valid_rows
        session.modified = True

        return render_template(
            "asset_import.html",
            valid_rows=valid_rows,
            invalid_rows=invalid_rows
        )

    return render_template("asset_import.html")


@app.route("/assets/import/confirm", methods=["POST"])
@login_required
def confirm_import():
    rows = session.pop("import_valid", [])

    if not rows:
        flash("No validated rows found to import.", "danger")
        return redirect(url_for("import_assets"))

    try:
        imported_count = 0
        for r in rows:

            if Asset.query.filter_by(asset_code=r["asset_code"]).first():
                continue

            if r.get("serial_no"):
                if Asset.query.filter_by(serial_no=r["serial_no"]).first():
                    continue

            asset = Asset(
                asset_code=r["asset_code"],
                asset_name=r["asset_name"],
                serial_no=r.get("serial_no"),
                plant=r.get("plant"),
                department=r.get("department"),
                location=r.get("location"),
                purchase_date=(
                    datetime.fromisoformat(r["purchase_date"]).date()
                    if r.get("purchase_date") else None
                ),
            )
            db.session.add(asset)
            imported_count += 1

        db.session.commit()

    except Exception as e:
        db.session.rollback()
        flash(f"Import failed: {str(e)}", "danger")
        return redirect(url_for("import_assets"))

    flash(f"{imported_count} assets imported successfully", "success")
    return redirect(url_for("asset_master"))


# -------------------------------------------------
# DOWNLOAD ASSET IMPORT TEMPLATE
# -------------------------------------------------
@app.route("/assets/import/template")
@login_required
def download_asset_template():
    wb = Workbook()
    ws = wb.active
    ws.title = "Asset Import Template"

    headers = [
        "Asset Code",
        "Asset Name",
        "Serial No",
        "Plant",
        "Department",
        "Location",
        "Purchase Date (DD/MM/YYYY)"
    ]

    ws.append(headers)

    for col in range(1, len(headers) + 1):
        ws.cell(row=1, column=col).font = ws.cell(row=1, column=col).font.copy(bold=True)
        ws.column_dimensions[chr(64 + col)].width = 22

    ws.append([
        "A001",
        "Hydraulic Press",
        "HP-8891",
        "Plant-1",
        "Production",
        "Line-3",
        "17/01/2026"
    ])

    output = BytesIO()
    wb.save(output)
    output.seek(0)

    return send_file(
        output,
        as_attachment=True,
        download_name="ACM_Asset_Import_Template.xlsx",
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )


# -------------------------------------------------
# AMC CREATE + ACTIVE LIST 
# -------------------------------------------------
@app.route("/amc", methods=["GET", "POST"])
@login_required
def amc_create():
    assets = Asset.query.order_by(Asset.id.asc()).all()

    if request.method == "POST":
        asset_id = request.form.get("asset_id")
        start_date = request.form.get("start_date")
        end_date = request.form.get("end_date")
        yearly_cost = request.form.get("yearly_cost")

        asset = Asset.query.get(asset_id)
        if not asset:
            flash("Invalid asset selected", "danger")
            return redirect(url_for("amc_create"))

        if asset.status == "Scrapped":
            flash("Cannot create AMC for scrapped asset", "danger")
            return redirect(url_for("amc_create"))

        existing_amc = AMC.query.filter_by(
            asset_id=asset.id,
            is_completed=False,
            is_cancelled=False
        ).first()

        if existing_amc:
            flash("Active AMC already exists for this asset", "danger")
            return redirect(url_for("amc_create"))

        try:
            start_date = datetime.strptime(start_date, "%Y-%m-%d").date()
            end_date = datetime.strptime(end_date, "%Y-%m-%d").date()
        except Exception:
            flash("Invalid date format", "danger")
            return redirect(url_for("amc_create"))

        if start_date >= end_date:
            flash("Start date must be before end date", "danger")
            return redirect(url_for("amc_create"))

        amc = AMC(
            asset_id=asset.id,
            start_date=start_date,
            end_date=end_date,
            yearly_cost=float(yearly_cost) if yearly_cost else None
        )

        try:
            db.session.add(amc)
            db.session.commit()

            flash("AMC created successfully", "success")
            return redirect(url_for("amc_view", amc_id=amc.id))

        except Exception:
            db.session.rollback()
            flash("AMC create failed. No data was saved.", "danger")
            return redirect(url_for("amc_create"))

    active_query = (
        AMC.query
        .filter_by(is_completed=False, is_cancelled=False)
        .order_by(AMC.end_date.asc())
    )

    page = request.args.get("page", 1, type=int)
    per_page = 15
    active_pagination = active_query.paginate(page=page, per_page=per_page, error_out=False)

    active_amcs = active_pagination.items

    amc_status_map = {}
    today = date.today()

    for amc in active_amcs:
        status_key, status_label = get_amc_status(amc)
        days_left = (amc.end_date - today).days

        amc_status_map[amc.id] = {
            "key": status_key,
            "label": status_label,
            "days_left": days_left
        }

    return render_template(
        "amc_create.html",
        assets=assets,
        active_amcs=active_amcs,
        amc_status_map=amc_status_map,
        pagination=active_pagination
    )


# -------------------------------------------------
# AMC VIEW
# -------------------------------------------------
@app.route("/amc/<int:amc_id>")
@login_required
def amc_view(amc_id):
    amc = AMC.query.get_or_404(amc_id)
    asset = amc.asset

    today = date.today()
    days_left = (amc.end_date - today).days

    status_key, status_label = get_amc_status(amc)

    asset_age = None
    if asset.purchase_date:
        asset_age = (today - asset.purchase_date).days

    event_total = sum(e.cost or 0 for e in amc.events)
    contract_value = amc.yearly_cost or 0
    total_spend = contract_value + event_total

    return render_template(
        "amc_view.html",
        amc=amc,
        asset=asset,
        status_key=status_key,
        status_label=status_label,
        asset_age=asset_age,
        days_left=days_left,
        contract_value=contract_value,
        event_total=event_total,
        total_spend=total_spend,
        today=today
    )


# -------------------------------------------------
# AMC EVENT CREATE 
# -------------------------------------------------
@app.route("/amc/<int:amc_id>/event", methods=["POST"])
@login_required
def amc_add_event(amc_id):
    amc = AMC.query.get_or_404(amc_id)

    if amc.is_completed or amc.is_cancelled:
        flash("Cannot add event to closed AMC", "danger")
        return redirect(url_for("amc_view", amc_id=amc.id))

    event_date = request.form.get("event_date")
    remarks = request.form.get("remarks")
    cost = request.form.get("cost")

    try:
        event_date = datetime.strptime(event_date, "%Y-%m-%d").date()
    except Exception:
        flash("Invalid event date", "danger")
        return redirect(url_for("amc_view", amc_id=amc.id))

    # cannot be future
    if event_date > date.today():
        flash("Event date cannot be in the future", "danger")
        return redirect(url_for("amc_view", amc_id=amc.id))

    # must fall between AMC period
    if event_date < amc.start_date or event_date > amc.end_date:
        flash("Event date must be between AMC start and end date", "danger")
        return redirect(url_for("amc_view", amc_id=amc.id))

    event = AMCEvent(
        amc_id=amc.id,
        event_date=event_date,
        remarks=remarks,
        cost=float(cost) if cost else None
    )

    try:
        db.session.add(event)
        db.session.commit()

        flash("AMC event added successfully", "success")
        return redirect(url_for("amc_view", amc_id=amc.id))

    except Exception:
        db.session.rollback()
        flash("AMC event save failed. No data was saved.", "danger")
        return redirect(url_for("amc_view", amc_id=amc.id))


# -------------------------------------------------
# AMC DOCUMENT UPLOAD
# -------------------------------------------------
@app.route("/amc/<int:amc_id>/document", methods=["POST"])
@login_required
def amc_upload_document(amc_id):
    amc = AMC.query.get_or_404(amc_id)

    if amc.is_completed or amc.is_cancelled:
        flash("Cannot upload document to closed AMC", "danger")
        return redirect(url_for("amc_view", amc_id=amc.id))

    file = request.files.get("file")
    doc_type = request.form.get("document_type")

    if doc_type == "Other":
        doc_type = request.form.get("other_document_type", "").strip()
        if not doc_type:
            flash("Please specify document type", "danger")
            return redirect(url_for("amc_view", amc_id=amc.id))

    if not file or not file.filename.lower().endswith(".pdf"):
        flash("Only PDF files are allowed", "danger")
        return redirect(url_for("amc_view", amc_id=amc.id))

    asset_code = amc.asset.asset_code
    month_year = datetime.today().strftime("%m-%Y")
    safe_doc_type = doc_type.replace(" ", "_")

    existing_count = AMCDocument.query.filter_by(
        amc_id=amc.id,
        document_type=doc_type
    ).count()

    seq = existing_count + 1

    stored_filename = f"{month_year}_AMC_{asset_code}_{safe_doc_type}_{seq}.pdf"
    file_path = os.path.join(AMC_DOC_DIR, stored_filename)

    try:
        file.save(file_path)

        doc = AMCDocument(
            amc_id=amc.id,
            document_type=doc_type,
            stored_filename=stored_filename,
            original_filename=file.filename
        )

        db.session.add(doc)
        db.session.commit()

        flash("Document uploaded successfully", "success")
        return redirect(url_for("amc_view", amc_id=amc.id))

    except Exception:
        db.session.rollback()

        try:
            if os.path.exists(file_path):
                os.remove(file_path)
        except Exception:
            pass

        flash("Document upload failed. No data was saved.", "danger")
        return redirect(url_for("amc_view", amc_id=amc.id))


# -------------------------------------------------
# AMC DOCUMENT DOWNLOAD
# -------------------------------------------------
@app.route("/amc/document/<int:doc_id>/download")
@login_required
def amc_download_document(doc_id):
    doc = AMCDocument.query.get_or_404(doc_id)
    file_path = os.path.join(AMC_DOC_DIR, doc.stored_filename)

    if not os.path.exists(file_path):
        flash("File not found on server", "danger")
        return redirect(url_for("amc_view", amc_id=doc.amc_id))

    return send_file(
        file_path,
        as_attachment=False,
        mimetype="application/pdf"
    )


# -------------------------------------------------
# AMC COMPLETE
# -------------------------------------------------
@app.route("/amc/<int:amc_id>/complete", methods=["GET", "POST"])
@login_required
def amc_complete(amc_id):
    amc = AMC.query.get_or_404(amc_id)

    if amc.is_completed or amc.is_cancelled:
        flash("AMC already closed", "warning")
        return redirect(url_for("amc_view", amc_id=amc.id))

    amc.is_completed = True
    amc.completed_on = date.today()

    try:
        db.session.commit()
        flash("AMC marked as completed", "success")
        return redirect(url_for("amc_create"))

    except Exception:
        db.session.rollback()
        flash("AMC completion failed. No changes were saved.", "danger")
        return redirect(url_for("amc_view", amc_id=amc.id))


# -------------------------------------------------
# AMC CANCEL
# -------------------------------------------------
@app.route("/amc/<int:amc_id>/cancel", methods=["GET", "POST"])
@login_required
def amc_cancel(amc_id):
    amc = AMC.query.get_or_404(amc_id)

    if amc.is_completed or amc.is_cancelled:
        flash("AMC already closed", "warning")
        return redirect(url_for("amc_view", amc_id=amc.id))

    amc.is_cancelled = True
    amc.completed_on = date.today()

    try:
        db.session.commit()
        flash("AMC marked as cancelled", "success")
        return redirect(url_for("amc_create"))

    except Exception:
        db.session.rollback()
        flash("AMC cancel failed. No changes were saved.", "danger")
        return redirect(url_for("amc_view", amc_id=amc.id))


# -------------------------------------------------
# CALIBRATION MAIN PAGE
# -------------------------------------------------
@app.route("/calibration")
@login_required
def calibration():
    assets = Asset.query.filter_by(status="Active").order_by(Asset.id.asc()).all()
    return render_template("calibration.html", assets=assets)


@app.route("/calibration/asset/<int:asset_id>")
@login_required
def calibration_asset_data(asset_id):
    asset = Asset.query.get_or_404(asset_id)

    if asset.status == "Scrapped":
        return jsonify({
            "success": False,
            "message": "Asset is scrapped"
        }), 400

    latest = (
        Calibration.query
        .filter_by(asset_id=asset.id)
        .order_by(
            Calibration.calibration_done_date.desc(),
            Calibration.created_on.desc()
        )
        .first()
    )

    today = date.today()

    asset_age = (
        (today - asset.purchase_date).days
        if asset.purchase_date else None
    )

    if latest:
        last_done = format_date_indian(latest.calibration_done_date)
        next_due = format_date_indian(latest.next_due_date)
        days_left = (latest.next_due_date - today).days
    else:
        last_done = next_due = days_left = "-"

    return {
        "asset": {
            "id": asset.id,
            "asset_code": asset.asset_code,
            "asset_name": asset.asset_name,
            "serial_no": asset.serial_no,
            "plant": asset.plant,
            "department": asset.department,
            "location": asset.location,
            "purchase_date": format_date_indian(asset.purchase_date),
        },
        "asset_age": asset_age,
        "last_done": last_done,
        "next_due": next_due,
        "days_left": days_left,
    }


@app.route("/calibration/save", methods=["POST"])
@login_required
def save_calibration():
    asset_id = request.form.get("asset_id")
    asset = Asset.query.get_or_404(asset_id)

    if asset.status == "Scrapped":
        return jsonify({
            "success": False,
            "message": "Cannot record calibration for scrapped asset"
        }), 400

    done_date = datetime.strptime(
        request.form["calibration_done_date"], "%Y-%m-%d"
    ).date()

    next_due = datetime.strptime(
        request.form["next_due_date"], "%Y-%m-%d"
    ).date()

    if next_due <= done_date:
        return jsonify({
            "success": False,
            "message": "Next calibration due date must be after calibration done date"
        }), 400

    if done_date > date.today():
        return jsonify({
            "success": False,
            "message": "Calibration done date cannot be in the future"
        }), 400

    try:
        cal = Calibration(
            asset_id=asset.id,
            calibration_done_date=done_date,
            next_due_date=next_due,
            cost=request.form.get("cost") or None,
            remarks=request.form.get("remarks"),
        )

        db.session.add(cal)
        db.session.flush()

        # --------------------
        # DOCUMENTS
        # --------------------
        files = request.files
        doc_index = 0

        while True:
            file = files.get(f"documents[{doc_index}][file]")
            if not file:
                break

            doc_type = request.form.get(f"documents[{doc_index}][type]")
            stored_filename = os.path.basename(
                request.form.get(f"documents[{doc_index}][filename]")
            )

            file_path = os.path.join(CALIBRATION_DOC_DIR, stored_filename)
            file.save(file_path)

            doc = CalibrationDocument(
                calibration_id=cal.id,
                document_type=doc_type,
                stored_filename=stored_filename,
                original_filename=file.filename
            )

            db.session.add(doc)
            doc_index += 1

        # --------------------
        # EVENTS 
        # --------------------
        idx = 0
        while f"event_date_{idx}" in request.form:
            ev_date_str = request.form.get(f"event_date_{idx}")
            ev_date = datetime.strptime(ev_date_str, "%Y-%m-%d").date()

            # cannot be future
            if ev_date > date.today():
                db.session.rollback()
                return jsonify({
                    "success": False,
                    "message": "Calibration event date cannot be in the future"
                }), 400

            # must be between done and next due
            if ev_date < done_date or ev_date > next_due:
                db.session.rollback()
                return jsonify({
                    "success": False,
                    "message": "Calibration event date must be between Done Date and Next Due Date"
                }), 400

            event = CalibrationEvent(
                calibration_id=cal.id,
                event_date=ev_date,
                remarks=request.form.get(f"event_remarks_{idx}"),
                cost=request.form.get(f"event_cost_{idx}") or None
            )
            db.session.add(event)
            idx += 1

        db.session.commit()

        return jsonify({
            "success": True,
            "message": "Calibration recorded successfully"
        })

    except Exception:
        db.session.rollback()
        return jsonify({
            "success": False,
            "message": "Calibration save failed. No data was recorded."
        }), 500


# -------------------------------------------------
# CALIBRATION DOCUMENT DOWNLOAD
# -------------------------------------------------
@app.route("/calibration/document/<int:doc_id>/download")
@login_required
def calibration_download_document(doc_id):
    doc = CalibrationDocument.query.get_or_404(doc_id)
    file_path = os.path.join(CALIBRATION_DOC_DIR, doc.stored_filename)

    if not os.path.exists(file_path):
        flash("File not found on server", "danger")
        return redirect(url_for("history_calibration_view", calibration_id=doc.calibration_id))

    return send_file(
        file_path,
        as_attachment=False,
        mimetype="application/pdf"
    )


# -------------------------------------------------
# HISTORY ROUTES (AMC) 
# -------------------------------------------------
@app.route("/history")
@login_required
def history():
    assets = Asset.query.order_by(Asset.asset_code.asc()).all()

    asset_id = request.args.get("asset_id")
    status = request.args.get("status")

    query = AMC.query.join(Asset)

    query = query.filter(
        (AMC.is_completed == True) | (AMC.is_cancelled == True)
    )

    if asset_id:
        query = query.filter(AMC.asset_id == asset_id)

    if status == "completed":
        query = query.filter(AMC.is_completed == True)

    if status == "cancelled":
        query = query.filter(AMC.is_cancelled == True)

    query = query.order_by(
        AMC.completed_on.desc().nullslast(),
        AMC.created_on.desc()
    )

    page = request.args.get("page", 1, type=int)
    per_page = 15
    pagination = query.paginate(page=page, per_page=per_page, error_out=False)

    return render_template(
        "history_amc_list.html",
        assets=assets,
        amcs=pagination.items,
        pagination=pagination,
        selected_asset=asset_id,
        selected_status=status
    )


@app.route("/history/amc/<int:amc_id>")
@login_required
def history_amc_view(amc_id):
    amc = AMC.query.get_or_404(amc_id)

    if not amc.is_completed and not amc.is_cancelled:
        flash("Active AMC cannot be viewed in history", "danger")
        return redirect(url_for("history"))

    asset = amc.asset

    event_total = sum(e.cost or 0 for e in amc.events)
    contract_value = amc.yearly_cost or 0
    total_spend = contract_value + event_total

    asset_age = None
    if asset.purchase_date:
        asset_age = (date.today() - asset.purchase_date).days

    return render_template(
        "history_amc_view.html",
        amc=amc,
        asset=asset,
        asset_age=asset_age,
        contract_value=contract_value,
        event_total=event_total,
        total_spend=total_spend
    )


@app.route("/history/amc/<int:amc_id>/export")
@login_required
def export_amc_history(amc_id):
    amc = AMC.query.get_or_404(amc_id)

    if not amc.is_completed and not amc.is_cancelled:
        flash("Active AMC cannot be exported", "danger")
        return redirect(url_for("history"))

    buffer = BytesIO()

    with zipfile.ZipFile(buffer, "w", zipfile.ZIP_DEFLATED) as z:

        summary = {
            "Asset Code": amc.asset.asset_code,
            "Asset Name": amc.asset.asset_name,
            "AMC Start": format_date_indian(amc.start_date),
            "AMC End": format_date_indian(amc.end_date),
            "Status": "Completed" if amc.is_completed else "Cancelled",
            "Contract Value": amc.yearly_cost or 0,
        }

        df_summary = pd.DataFrame([summary])
        summary_xlsx = BytesIO()
        df_summary.to_excel(summary_xlsx, index=False)
        z.writestr("summary/amc_summary.xlsx", summary_xlsx.getvalue())

        events = [{
            "Date": format_date_indian(e.event_date),
            "Remarks": e.remarks,
            "Cost": e.cost
        } for e in amc.events]

        df_events = pd.DataFrame(events)
        events_xlsx = BytesIO()
        df_events.to_excel(events_xlsx, index=False)
        z.writestr("events/amc_events.xlsx", events_xlsx.getvalue())

        for d in amc.documents:
            path = os.path.join(AMC_DOC_DIR, d.stored_filename)
            if os.path.exists(path):
                z.write(path, arcname=f"documents/{d.stored_filename}")

    buffer.seek(0)

    filename = f"AMC_{amc.asset.asset_code}_{amc.id}.zip"

    return send_file(
        buffer,
        as_attachment=True,
        download_name=filename,
        mimetype="application/zip"
    )


# -------------------------------------------------
# HISTORY ROUTES (Calibration) 
# -------------------------------------------------
@app.route("/history/calibration")
@login_required
def history_calibration_list():
    assets = Asset.query.order_by(Asset.asset_code.asc()).all()
    asset_id = request.args.get("asset_id")

    query = Calibration.query.join(Asset)

    if asset_id:
        query = query.filter(Calibration.asset_id == asset_id)

    query = query.order_by(
        Calibration.calibration_done_date.desc(),
        Calibration.next_due_date.desc(),
        Calibration.created_on.desc()
    )

    page = request.args.get("page", 1, type=int)
    per_page = 15
    pagination = query.paginate(page=page, per_page=per_page, error_out=False)

    calibrations = pagination.items
    today = date.today()

    all_for_status = query.all()

    latest_per_asset = {}
    for c in all_for_status:
        current = latest_per_asset.get(c.asset_id)

        if not current:
            latest_per_asset[c.asset_id] = c
            continue

        if c.calibration_done_date > current.calibration_done_date:
            latest_per_asset[c.asset_id] = c
        elif c.calibration_done_date == current.calibration_done_date:
            if c.next_due_date > current.next_due_date:
                latest_per_asset[c.asset_id] = c
            elif c.next_due_date == current.next_due_date:
                if c.created_on > current.created_on:
                    latest_per_asset[c.asset_id] = c

    latest_per_asset = {k: v.id for k, v in latest_per_asset.items()}

    days_left_map = {}
    status_map = {}

    for c in calibrations:
        days_left = (c.next_due_date - today).days
        days_left_map[c.id] = days_left

        if latest_per_asset.get(c.asset_id) != c.id:
            status_map[c.id] = "superseded"
        else:
            if days_left < 0:
                status_map[c.id] = "overdue"
            else:
                status_map[c.id] = "valid"

    return render_template(
        "history_calibration_list.html",
        assets=assets,
        calibrations=calibrations,
        pagination=pagination,
        days_left_map=days_left_map,
        status_map=status_map,
        selected_asset=asset_id
    )


@app.route("/history/calibration/<int:calibration_id>")
@login_required
def history_calibration_view(calibration_id):
    calibration = Calibration.query.get_or_404(calibration_id)
    asset = calibration.asset

    today = date.today()
    days_left = (calibration.next_due_date - today).days

    asset_age = None
    if asset.purchase_date:
        asset_age = (today - asset.purchase_date).days

    latest = (
        Calibration.query
        .filter_by(asset_id=asset.id)
        .order_by(
            Calibration.calibration_done_date.desc(),
            Calibration.next_due_date.desc(),
            Calibration.created_on.desc()
        )
        .first()
    )

    if latest and latest.id != calibration.id:
        status = "superseded"
    elif days_left < 0:
        status = "overdue"
    elif days_left == 0:
        status = "due_today"
    else:
        status = "valid"

    event_total = sum(e.cost or 0 for e in calibration.events)
    calibration_cost = calibration.cost or 0
    total_spend = calibration_cost + event_total

    return render_template(
        "history_calibration_view.html",
        calibration=calibration,
        asset=asset,
        asset_age=asset_age,
        days_left=days_left,
        status=status,
        event_total=event_total,
        calibration_cost=calibration_cost,
        total_spend=total_spend
    )


@app.route("/history/calibration/<int:calibration_id>/export")
@login_required
def export_calibration_history(calibration_id):
    calibration = Calibration.query.get_or_404(calibration_id)

    buffer = BytesIO()

    with zipfile.ZipFile(buffer, "w", zipfile.ZIP_DEFLATED) as z:

        summary = {
            "Asset Code": calibration.asset.asset_code,
            "Asset Name": calibration.asset.asset_name,
            "Calibration Date": format_date_indian(calibration.calibration_done_date),
            "Next Due Date": format_date_indian(calibration.next_due_date),
            "Cost": calibration.cost or 0,
            "Remarks": calibration.remarks or "-"
        }

        df_summary = pd.DataFrame([summary])
        summary_xlsx = BytesIO()
        df_summary.to_excel(summary_xlsx, index=False)
        z.writestr("summary/calibration_summary.xlsx", summary_xlsx.getvalue())

        events = [{
            "Date": format_date_indian(e.event_date),
            "Remarks": e.remarks,
            "Cost": e.cost
        } for e in calibration.events]

        df_events = pd.DataFrame(events)
        events_xlsx = BytesIO()
        df_events.to_excel(events_xlsx, index=False)
        z.writestr("events/calibration_events.xlsx", events_xlsx.getvalue())

        for d in calibration.documents:
            path = os.path.join(CALIBRATION_DOC_DIR, d.stored_filename)
            if os.path.exists(path):
                z.write(path, arcname=f"documents/{d.stored_filename}")

    buffer.seek(0)

    filename = f"Calibration_{calibration.asset.asset_code}_{calibration.id}.zip"

    return send_file(
        buffer,
        as_attachment=True,
        download_name=filename,
        mimetype="application/zip"
    )


# -------------------------------------------------
# SCRAP ASSET
# -------------------------------------------------
@app.route("/scrap", methods=["GET", "POST"])
@login_required
@role_required(["developer", "purchase"])
def scrap_asset():
    assets = (
        Asset.query
        .filter_by(status="Active")
        .order_by(Asset.asset_code.asc())
        .all()
    )

    if request.method == "POST":
        asset_id = request.form.get("asset_id")
        scrap_date = request.form.get("scrap_date")
        approved_by = request.form.get("approved_by")
        file = request.files.get("scrap_note")

        asset = Asset.query.get_or_404(asset_id)

        if asset.status == "Scrapped":
            flash("Asset is already scrapped", "danger")
            return redirect(url_for("scrap_asset"))

        if not scrap_date or not approved_by or not file:
            flash("All fields are mandatory", "danger")
            return redirect(url_for("scrap_asset"))

        if not file.filename.lower().endswith(".pdf"):
            flash("Scrap note must be a PDF file", "danger")
            return redirect(url_for("scrap_asset"))

        scrap_date = datetime.strptime(scrap_date, "%Y-%m-%d").date()

        filename = f"{scrap_date}_SCRAP_{asset.asset_code}.pdf"
        file_path = os.path.join(SCRAP_DOC_DIR, filename)
        file.save(file_path)

        try:
            scrap = AssetScrap(
                asset_id=asset.id,
                scrap_date=scrap_date,
                approved_by=approved_by.strip(),
                scrap_note_filename=filename,
                original_filename=file.filename
            )
            db.session.add(scrap)

            asset.status = "Scrapped"

            AMC.query.filter(
                AMC.asset_id == asset.id,
                AMC.is_completed == False,
                AMC.is_cancelled == False
            ).update({
                AMC.is_cancelled: True,
                AMC.completed_on: scrap_date
            })

            db.session.commit()

            flash("Asset scrapped successfully", "success")
            return redirect(url_for("scrap_asset"))

        except Exception:
            db.session.rollback()
            flash("Scrap failed. No changes were saved.", "danger")
            return redirect(url_for("scrap_asset"))

    return render_template("scrap_asset.html", assets=assets)


# -------------------------------------------------
# SCRAP HISTORY 
# -------------------------------------------------
@app.route("/history/scrap")
@login_required
def history_scrap_list():
    query = (
        AssetScrap.query
        .join(Asset)
        .order_by(AssetScrap.scrap_date.desc())
    )

    page = request.args.get("page", 1, type=int)
    per_page = 15
    pagination = query.paginate(page=page, per_page=per_page, error_out=False)

    return render_template(
        "history_scrap_list.html",
        scraps=pagination.items,
        pagination=pagination
    )


@app.route("/scrap/document/<int:scrap_id>/download")
@login_required
def scrap_download(scrap_id):
    scrap = AssetScrap.query.get_or_404(scrap_id)
    path = os.path.join(SCRAP_DOC_DIR, scrap.scrap_note_filename)

    if not os.path.exists(path):
        flash("Scrap note file not found", "danger")
        return redirect(url_for("history_scrap_list"))

    return send_file(path, as_attachment=False, mimetype="application/pdf")


# -------------------------------
# REMINDER ROUTES
# -------------------------------
@app.route("/reminders")
@login_required
def reminders():
    today = date.today()
    reminders = []

    # AMC reminders
    active_amcs = (
        AMC.query
        .join(Asset)
        .filter(
            Asset.status == "Active",
            AMC.is_completed == False,
            AMC.is_cancelled == False
        )
        .all()
    )

    for amc in active_amcs:
        days_left = (amc.end_date - today).days

        if days_left < 0:
            rule = "overdue"
            severity = "overdue"
        elif 0 <= days_left <= 7:
            rule = "due_soon"
            severity = "due_soon"
        elif 8 <= days_left <= 15:
            rule = "upcoming"
            severity = "upcoming"
        else:
            continue

        if not is_acknowledged("AMC", amc.id, rule):
            reminders.append({
                "type": "AMC",
                "severity": severity,
                "rule": rule,
                "asset_id": amc.asset.id,
                "asset_code": amc.asset.asset_code,
                "asset_name": amc.asset.asset_name,
                "days": days_left,
                "due_date": amc.end_date,
                "source_id": amc.id
            })

    active_assets = Asset.query.filter_by(status="Active").all()

    for asset in active_assets:
        latest_cal = (
            Calibration.query
            .filter_by(asset_id=asset.id)
            .order_by(
                Calibration.calibration_done_date.desc(),
                Calibration.created_on.desc()
            )
            .first()
        )

        if not latest_cal:
            continue

        days_left = (latest_cal.next_due_date - today).days

        if days_left < 0:
            rule = "overdue"
            severity = "overdue"
        elif 0 <= days_left <= 7:
            rule = "due_soon"
            severity = "due_soon"
        elif 8 <= days_left <= 15:
            rule = "upcoming"
            severity = "upcoming"
        else:
            continue

        if not is_acknowledged("Calibration", latest_cal.id, rule):
            reminders.append({
                "type": "Calibration",
                "severity": severity,
                "rule": rule,
                "asset_id": asset.id,
                "asset_code": asset.asset_code,
                "asset_name": asset.asset_name,
                "days": days_left,
                "due_date": latest_cal.next_due_date,
                "source_id": latest_cal.id
            })

    severity_order = {
        "overdue": 0,
        "due_soon": 1,
        "upcoming": 2
    }

    reminders.sort(
        key=lambda r: (
            severity_order[r["severity"]],
            abs(r["days"])
        )
    )

    return render_template("reminders.html", reminders=reminders)


@app.route("/reminders/acknowledge", methods=["POST"])
@login_required
def acknowledge_reminder():
    source_type = request.json.get("source_type")
    source_id = request.json.get("source_id")
    rule = request.json.get("rule")

    if not source_type or not source_id or not rule:
        return jsonify({
            "success": False,
            "message": "Invalid reminder data"
        }), 400

    existing = ReminderAck.query.filter_by(
        source_type=source_type,
        source_id=source_id,
        rule=rule
    ).first()

    if existing:
        return jsonify({"success": True})

    if source_type == "AMC":
        if not AMC.query.get(source_id):
            return jsonify({"success": False, "message": "Invalid AMC"}), 400

    elif source_type == "Calibration":
        if not Calibration.query.get(source_id):
            return jsonify({"success": False, "message": "Invalid Calibration"}), 400

    ack = ReminderAck(
        source_type=source_type,
        source_id=source_id,
        rule=rule,
        asset_id=request.json.get("asset_id"),
        acknowledged_by=session.get("user_id")
    )

    try:
        db.session.add(ack)
        db.session.commit()
        return jsonify({"success": True})

    except Exception:
        db.session.rollback()
        return jsonify({
            "success": False,
            "message": "Acknowledgement failed. Please refresh and try again."
        }), 500


# -------------------------------------------------
# BACKUP
# -------------------------------------------------
@app.route("/backup")
@login_required
def backup():
    if not os.path.exists(DATA_DIR):
        flash("Data directory not found. Backup failed.", "danger")
        return redirect(url_for("reminders"))

    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    backup_name = f"ACM_Backup_{timestamp}.zip"

    buffer = BytesIO()

    with zipfile.ZipFile(buffer, "w", zipfile.ZIP_DEFLATED) as zipf:
        for root, dirs, files in os.walk(DATA_DIR):
            for file in files:
                full_path = os.path.join(root, file)
                arcname = os.path.relpath(full_path, BASE_DIR)
                zipf.write(full_path, arcname)

    buffer.seek(0)

    return send_file(
        buffer,
        as_attachment=True,
        download_name=backup_name,
        mimetype="application/zip"
    )


# -------------------------------------------------
# CHANGE PASSWORD
# -------------------------------------------------
@app.route("/change-password", methods=["GET", "POST"])
@login_required
def change_password():
    username = session.get("username")

    if request.method == "POST":
        current = request.form.get("current_password")
        new = request.form.get("new_password")
        confirm = request.form.get("confirm_password")

        if not current or not new or not confirm:
            flash("All fields are required.", "danger")
            return redirect(url_for("change_password"))

        if new != confirm:
            flash("New passwords do not match.", "danger")
            return redirect(url_for("change_password"))

        user = User.query.filter_by(username=username).first()

        if not user or not check_password_hash(user.password_hash, current):
            flash("Current password is incorrect.", "danger")
            return redirect(url_for("change_password"))

        try:
            user.password_hash = generate_password_hash(new)
            db.session.commit()

            flash("Password changed successfully.", "success")
            return redirect(url_for("reminders"))

        except Exception:
            db.session.rollback()
            flash("Password change failed. Try again.", "danger")
            return redirect(url_for("change_password"))

    return render_template("change_password.html")


# -------------------------------------------------
# PASSWORD RESET
# -------------------------------------------------
@app.route("/reset-password", methods=["GET", "POST"])
@login_required
def reset_password():
    role = session.get("role")

    if role != "developer":
        flash("Access denied.", "danger")
        return redirect(url_for("reminders"))

    users = User.query.all()

    if request.method == "POST":
        username = request.form.get("username")
        new_password = request.form.get("new_password")

        if not username or not new_password:
            flash("All fields are required.", "danger")
            return redirect(url_for("reset_password"))

        user = User.query.filter_by(username=username).first()

        if not user:
            flash("User not found.", "danger")
            return redirect(url_for("reset_password"))

        try:
            user.password_hash = generate_password_hash(new_password)
            db.session.commit()

            flash(f"Password reset for {username}.", "success")
            return redirect(url_for("reset_password"))

        except Exception:
            db.session.rollback()
            flash("Password reset failed. No changes were saved.", "danger")
            return redirect(url_for("reset_password"))

    return render_template("reset_password.html", users=users)


# -------------------------------------------------
# RUN
# -------------------------------------------------
if __name__ == "__main__":
    app.run(host="0.0.0.0", port=5001, debug=False)
